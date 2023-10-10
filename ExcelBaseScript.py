import openpyxl
from openpyxl import workbook

LoginValidUserNameFile = openpyxl.load_workbook('C:\\Users\\owner\\PycharmProjects\\Omni-CAL\\CAL\\Test\\TestData\\LoginInvalidUserNamePassword.xlsx')
print (type(LoginValidUserNameFile))
#print(LoginValidUserNameFile.)
#print(wb['LoginValidUserNameFile'])
print(LoginValidUserNameFile.get_sheet_names())
sheet1 = LoginValidUserNameFile.get_sheet_by_name('Sheet1')
print(sheet1['A2'].value)
a=sheet1.cell(row=3, column=2).value
print(a)
multiple_cells = sheet1['A1':'B5']
for row in multiple_cells:
    for cell in row:
        print(cell.value)

all_rows = [sheet1.rows]
print(all_rows[:])
#print(all_rows[:])

all_columns = (sheet1.columns)
print(all_columns[:])
# wb = openpyxl.load_workbook('C:\\Users\\owner\\PycharmProjects\\Omni-CAL\\CAL\\Test\\TestData\\LoginInvalidUserNamePassword.xlsx')
# Sheetname =  wb.worksheets[0]
# print(Sheetname)

